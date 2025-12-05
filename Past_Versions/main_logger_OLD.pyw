# Thermal Temp Controller Logger - GUI + Colored Excel + Live Graph in separate window
# LUX Dynamics - Kailani Alarcon

import time
import csv
import os
import math
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox
from typing import Dict, List, Tuple

from tc08_interface import TC08Interface  # your existing TC-08 interface

# Excel support (for pretty colored columns)
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Border, Side
    HAVE_OPENPYXL = True
except ImportError:
    Workbook = None
    PatternFill = None
    Border = None
    Side = None
    HAVE_OPENPYXL = False

# Arduino support (pyserial)
try:
    import serial
    HAVE_SERIAL = True
except ImportError:
    serial = None
    HAVE_SERIAL = False

# Preferred folder for logs
OUTPUT_FOLDER = r"Z:\ENGINEERING\Product Development\Thermal Testing 2025"


# ---------------- Helpers: filenames, Excel, formatting ---------------- #

def get_unique_csv_path(folder: str, base_name: str) -> str:
    """
    Return a unique CSV path in 'folder' based on 'base_name'.
    """
    path = os.path.join(folder, base_name + ".csv")
    if not os.path.exists(path):
        return path
    i = 1
    while True:
        alt = os.path.join(folder, f"{base_name}_{i}.csv")
        if not os.path.exists(alt):
            return alt
        i += 1


def resolve_output_folder() -> str:
    """
    Use Z: folder if available; otherwise local ./logs.
    """
    if os.path.isdir(OUTPUT_FOLDER):
        return OUTPUT_FOLDER
    fallback = os.path.join(os.getcwd(), "logs")
    os.makedirs(fallback, exist_ok=True)
    return fallback


def apply_column_colors(ws):
    """
    Color each column from the header row downward with a unique pale solid color
    and add bolder grid lines so columns stand out.
    """
    if not HAVE_OPENPYXL or PatternFill is None:
        return

    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value == "timestamp":
                header_row_idx = cell.row
                break
        if header_row_idx is not None:
            break
    if header_row_idx is None:
        return

    header_cells = list(ws[header_row_idx])
    num_cols = len(header_cells)
    if num_cols == 0:
        return

    palette = [
        "FFCCCC", "FFE5CC", "FFF2CC", "E5FFCC",
        "CCFFFF", "CCE5FF", "E5CCFF", "FFCCF2", "E6E6FA",
    ]

    bold_border = Border(
        left=Side(style="medium", color="000000"),
        right=Side(style="medium", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col_idx, cell in enumerate(header_cells, start=1):
        color_hex = palette[(col_idx - 1) % len(palette)]
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

        for row_idx in range(header_row_idx, ws.max_row + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.fill = fill
            c.border = bold_border


def create_colored_excel(csv_path: str):
    """
    Read the CSV and create a colored Excel file (.xlsx) with each column
    tinted with a solid pale color and bold column borders.
    """
    if not HAVE_OPENPYXL or Workbook is None:
        print("openpyxl not available → skipping colored Excel export.")
        return

    xlsx_path = os.path.splitext(csv_path)[0] + ".xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "TC08 Log"

    with open(csv_path, newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    apply_column_colors(ws)
    wb.save(xlsx_path)
    print(f"Colored Excel copy saved as:\n  {xlsx_path}")


def fmt_val(val):
    """
    Format numeric value to 2 decimal places for CSV/Excel.
    Returns '' for NaN / None so it doesn't blow up.
    """
    try:
        if val is None:
            return ""
        if isinstance(val, float) and math.isnan(val):
            return ""
        return f"{float(val):.2f}"
    except (TypeError, ValueError):
        return ""


# ---------------- Arduino Interface ---------------- #

class ArduinoInterface:
    """
    Arduino prints lines like:
      TEMP:25.30,HOLD:53.60,PWM:255
    and accepts commands like:
      SET:25.0   (set holdC to 25°C)
    """

    def __init__(self, port: str, baudrate: int = 9600):
        if not HAVE_SERIAL:
            raise RuntimeError(
                "pyserial not installed; cannot use ArduinoInterface. Get Kailani."
            )
        self.ser = serial.Serial(port, baudrate=baudrate, timeout=0.1)
        time.sleep(2.0)
        self.ser.reset_input_buffer()
        self.latest_temp = None
        self.latest_hold = None
        self.latest_pwm = None

    def set_hold(self, temp_c: float):
        cmd = f"SET:{temp_c:.2f}\n"
        try:
            self.ser.write(cmd.encode("ascii"))
        except Exception:
            pass

    def poll(self):
        line = None
        try:
            while self.ser.in_waiting:
                raw = self.ser.readline()
                if not raw:
                    break
                line = raw.decode("ascii", errors="ignore").strip()
        except Exception:
            return self.latest_temp, self.latest_hold, self.latest_pwm

        if not line:
            return self.latest_temp, self.latest_hold, self.latest_pwm

        try:
            if "TEMP:" in line:
                parts = [p.strip() for p in line.split(",")]
                for p in parts:
                    if p.startswith("TEMP:"):
                        self.latest_temp = float(p.split("TEMP:")[1])
                    elif p.startswith("HOLD:"):
                        self.latest_hold = float(p.split("HOLD:")[1])
                    elif p.startswith("PWM:"):
                        self.latest_pwm = float(p.split("PWM:")[1])
            else:
                self.latest_temp = float(line)
        except ValueError:
            pass

        return self.latest_temp, self.latest_hold, self.latest_pwm

    def close(self):
        try:
            self.ser.close()
        except Exception:
            pass


# ---------------- Live Graph Window ---------------- #

class LiveGraphWindow(tk.Toplevel):
    """
    Separate window that shows a live graph of time vs temperature
    for all active TC-08 channels.

    - Zoom in/out with +/- buttons (time window, minutes).
    - Scroll through history with a styled slider (Earlier → Later).
    - Auto y-axis scaling.
    - Small y-axis with ticks + numeric labels.
    - Hover readout at bottom showing time + channel values.
    - Channel visibility checkboxes.
    """

    def __init__(self, master):
        super().__init__(master)
        self.title("Live Temperature Graph")
        self.geometry("950x500")

        self.history: Dict[int, Dict[str, List[float]]] = {}
        self.active_channels: List[Tuple[int, str]] = []
        self.window_sec = 300.0  # default 5 minutes
        self.max_points = 2000
        self.graph_colors = [
            "blue", "red", "green", "purple",
            "orange", "brown", "magenta", "cyan"
        ]
        self.pan_var = tk.DoubleVar(value=0.0)

        # axis / plot geometry state for hover
        self.plot_left = None
        self.plot_right = None
        self.plot_top = None
        self.plot_bottom = None
        self.tmin = None
        self.tmax = None
        self.vmin = None
        self.vmax = None

        # channel visibility
        self.channel_visibility: Dict[int, tk.BooleanVar] = {}

        self._build_ui()
        self._update_window_label()

    def _build_ui(self):
        controls = ttk.Frame(self, padding=8)
        controls.pack(fill="x")

        self.window_label_var = tk.StringVar()
        ttk.Button(controls, text="Zoom -", command=self.zoom_out).pack(side="left")
        ttk.Button(controls, text="Zoom +", command=self.zoom_in).pack(side="left", padx=(2, 8))
        ttk.Label(controls, textvariable=self.window_label_var).pack(side="left")

        # prettier slider
        style = ttk.Style(self)
        style.configure(
            "Pan.Horizontal.TScale",
            troughcolor="#e5e5e5",
        )

        slider_frame = ttk.Frame(controls)
        slider_frame.pack(side="right")
        ttk.Label(slider_frame, text="Earlier").pack(side="left", padx=(0, 4))
        self.pan_scale = ttk.Scale(
            slider_frame,
            from_=0.0,
            to=100.0,
            orient="horizontal",
            variable=self.pan_var,
            command=lambda v: self.redraw(),
            style="Pan.Horizontal.TScale",
            length=260,
        )
        self.pan_scale.pack(side="left")
        ttk.Label(slider_frame, text="Later").pack(side="left", padx=(4, 0))

        # channel toggles
        self.toggle_frame = ttk.Frame(self, padding=(8, 0))
        self.toggle_frame.pack(fill="x", anchor="w")

        self.canvas = tk.Canvas(self, bg="white")
        self.canvas.pack(fill="both", expand=True, padx=8, pady=(0, 4))
        self.canvas.bind("<Motion>", self.on_mouse_move)
        self.canvas.bind("<Leave>", lambda e: self.hover_label_var.set(""))

        # hover readout
        self.hover_label_var = tk.StringVar(value="Hover over plot for values")
        self.hover_label = ttk.Label(self, textvariable=self.hover_label_var, padding=(8, 2))
        self.hover_label.pack(fill="x", side="bottom", anchor="w")

        self.protocol("WM_DELETE_WINDOW", self.on_close)

    def on_close(self):
        """Just close this window; logger keeps running."""
        self.destroy()

    def _update_window_label(self):
        if self.window_sec is None:
            text = "Window: full"
        else:
            text = f"Window: {self.window_sec / 60.0:.2f} min"
        self.window_label_var.set(text)

    def set_channels(self, active_channels: List[Tuple[int, str]]):
        self.active_channels = list(active_channels)
        self.history.clear()
        self.refresh_channel_toggles()

    def refresh_channel_toggles(self):
        # clear old
        for child in self.toggle_frame.winfo_children():
            child.destroy()
        # build new
        for ch, name in self.active_channels:
            var = self.channel_visibility.get(ch)
            if var is None:
                var = tk.BooleanVar(value=True)
                self.channel_visibility[ch] = var
            cb = ttk.Checkbutton(
                self.toggle_frame,
                text=name,
                variable=var,
                command=self.redraw
            )
            cb.pack(side="left", padx=(0, 8))

    def add_sample(self, elapsed: float, temps: Dict[int, float]):
        # append new samples
        for ch, _name in self.active_channels:
            val = temps.get(ch, float("nan"))
            try:
                if val is None or math.isnan(float(val)):
                    continue
            except (TypeError, ValueError):
                continue
            if ch not in self.history:
                self.history[ch] = {"t": [], "v": []}
            self.history[ch]["t"].append(float(elapsed))
            self.history[ch]["v"].append(float(val))
            if len(self.history[ch]["t"]) > self.max_points:
                self.history[ch]["t"] = self.history[ch]["t"][-self.max_points:]
                self.history[ch]["v"] = self.history[ch]["v"][-self.max_points:]
        self.redraw()

    def zoom_in(self):
        if self.window_sec is None:
            self.window_sec = 300.0
        self.window_sec = max(5.0, self.window_sec / 2.0)
        self._update_window_label()
        self.redraw()

    def zoom_out(self):
        if not self.history:
            return
        all_times: List[float] = []
        for h in self.history.values():
            all_times.extend(h["t"])
        if not all_times:
            return
        total_span = max(all_times) - min(all_times)
        if total_span <= 0:
            return
        if self.window_sec is None:
            self.window_sec = total_span
        self.window_sec *= 2.0
        if self.window_sec >= total_span:
            self.window_sec = None  # full
        self._update_window_label()
        self.redraw()

    def redraw(self):
        # reset hover geometry
        self.plot_left = self.plot_right = self.plot_top = self.plot_bottom = None
        self.tmin = self.tmax = self.vmin = self.vmax = None

        if not self.history:
            self.canvas.delete("all")
            return

        # collect global time/value range
        all_times: List[float] = []
        all_vals: List[float] = []
        for h in self.history.values():
            all_times.extend(h["t"])
            all_vals.extend(h["v"])

        if len(all_times) < 2 or not all_vals:
            self.canvas.delete("all")
            return

        global_tmin = min(all_times)
        global_tmax = max(all_times)

        # determine displayed tmin/tmax
        if self.window_sec is None:
            tmin = global_tmin
            tmax = global_tmax
        else:
            window = self.window_sec
            span = max(global_tmax - global_tmin, 1e-6)
            window = min(window, span)
            start_min = global_tmin
            start_max = global_tmax - window
            if start_max <= start_min:
                tmin = global_tmin
            else:
                frac = max(0.0, min(1.0, self.pan_var.get() / 100.0))
                tmin = start_min + frac * (start_max - start_min)
            tmax = tmin + window
        if tmax <= tmin:
            tmax = tmin + 1.0

        # y-limits auto
        vmin = min(all_vals)
        vmax = max(all_vals)
        if vmax <= vmin:
            vmax = vmin + 1.0

        w = self.canvas.winfo_width() or 400
        h = self.canvas.winfo_height() or 300

        legend_width = 130
        margin_right = 40   # give a bit more room right side
        margin_top = 20
        margin_bottom = 30
        margin_left_extra = 30  # space for y labels

        plot_left = legend_width + margin_left_extra
        plot_right = w - margin_right
        plot_top = margin_top
        plot_bottom = h - margin_bottom

        if plot_right <= plot_left + 10:
            self.canvas.delete("all")
            return

        # store geometry for hover
        self.plot_left = plot_left
        self.plot_right = plot_right
        self.plot_top = plot_top
        self.plot_bottom = plot_bottom
        self.tmin = tmin
        self.tmax = tmax
        self.vmin = vmin
        self.vmax = vmax

        self.canvas.delete("all")

        # grid + x & y tick labels
        n_x = 5
        n_y = 5
        time_span = tmax - tmin

        # vertical grid + x labels (time)
        for i in range(n_x + 1):
            gx = plot_left + i * (plot_right - plot_left) / n_x
            self.canvas.create_line(
                gx, plot_top, gx, plot_bottom,
                fill="#e0e0e0"
            )
            t_here = tmin + time_span * i / n_x
            label = f"{t_here / 60.0:.1f}"  # minutes
            self.canvas.create_text(
                gx, plot_bottom + 12,
                text=label,
                anchor="n"
            )

        # horizontal grid + y labels (temp)
        for j in range(n_y + 1):
            gy = plot_top + j * (plot_bottom - plot_top) / n_y
            self.canvas.create_line(
                plot_left, gy, plot_right, gy,
                fill="#e0e0e0"
            )
            # small y-axis with numeric labels
            v_here = vmax - (vmax - vmin) * j / n_y
            self.canvas.create_line(
                plot_left - 4, gy, plot_left, gy,
                fill="black"
            )
            self.canvas.create_text(
                plot_left - 6, gy,
                text=f"{v_here:.1f}",
                anchor="e"
            )

        # axes
        self.canvas.create_line(
            plot_left, plot_bottom, plot_right, plot_bottom,
            fill="black", width=2
        )
        self.canvas.create_line(
            plot_left, plot_top, plot_left, plot_bottom,
            fill="black", width=2
        )

        # plot each visible channel
        for idx, (ch, name) in enumerate(self.active_channels):
            var = self.channel_visibility.get(ch)
            if var is not None and not var.get():
                continue

            if ch not in self.history:
                continue
            t_list = self.history[ch]["t"]
            v_list = self.history[ch]["v"]
            if len(t_list) < 2:
                continue

            coords: List[float] = []
            for t_val, v_val in zip(t_list, v_list):
                if t_val < tmin or t_val > tmax:
                    continue
                x = plot_left + (t_val - tmin) / (tmax - tmin) * (plot_right - plot_left)
                y = plot_bottom - (v_val - vmin) / (vmax - vmin) * (plot_bottom - plot_top)
                coords.extend((x, y))

            if len(coords) < 4:
                continue

            color = self.graph_colors[idx % len(self.graph_colors)]
            self.canvas.create_line(*coords, fill=color, width=2)

        # x-axis label
        window_min = (tmax - tmin) / 60.0
        self.canvas.create_text(
            (plot_left + plot_right) / 2,
            h - 10,
            text=f"Time (min) – window ≈ {window_min:.2f} min",
            anchor="center"
        )

        # legend (only visible channels)
        legend_x = 10
        legend_y = 25
        for idx, (ch, name) in enumerate(self.active_channels):
            var = self.channel_visibility.get(ch)
            if var is not None and not var.get():
                continue
            color = self.graph_colors[idx % len(self.graph_colors)]
            self.canvas.create_rectangle(
                legend_x,
                legend_y - 5,
                legend_x + 20,
                legend_y + 5,
                fill=color,
                outline=color
            )
            self.canvas.create_text(
                legend_x + 25,
                legend_y,
                text=name,
                anchor="w"
            )
            legend_y += 18

    def on_mouse_move(self, event):
        """Update hover label with nearest time + values of visible channels."""
        if not self.history:
            self.hover_label_var.set("")
            return
        if (
            self.plot_left is None or self.plot_right is None or
            self.plot_top is None or self.plot_bottom is None or
            self.tmin is None or self.tmax is None
        ):
            self.hover_label_var.set("")
            return

        x = event.x
        y = event.y
        if x < self.plot_left or x > self.plot_right or y < self.plot_top or y > self.plot_bottom:
            self.hover_label_var.set("")
            return

        if self.plot_right == self.plot_left or self.tmax == self.tmin:
            self.hover_label_var.set("")
            return

        t_current = self.tmin + (x - self.plot_left) / (self.plot_right - self.plot_left) * (self.tmax - self.tmin)

        info = []
        for ch, name in self.active_channels:
            var = self.channel_visibility.get(ch)
            if var is not None and not var.get():
                continue
            if ch not in self.history:
                continue
            t_list = self.history[ch]["t"]
            v_list = self.history[ch]["v"]
            if not t_list:
                continue

            best_idx = None
            best_dt = None
            for idx, t_val in enumerate(t_list):
                if t_val < self.tmin or t_val > self.tmax:
                    continue
                dt = abs(t_val - t_current)
                if best_idx is None or dt < best_dt:
                    best_idx = idx
                    best_dt = dt
            if best_idx is None:
                continue
            t_val = t_list[best_idx]
            v_val = v_list[best_idx]
            info.append((name, t_val, v_val))

        if not info:
            self.hover_label_var.set("")
            return

        t_display = info[0][1] / 60.0  # minutes
        parts = [f"t={t_display:.2f} min"]
        for name, _, v in info:
            parts.append(f"{name}={v:.2f}°C")
        self.hover_label_var.set(" | ".join(parts))


# ---------------- Main GUI App ---------------- #

class ThermalLoggerApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("LUX Thermal Thermal Logger")
        self.geometry("900x600")

        self.logger = None
        self.csv_file = None
        self.csv_writer = None
        self.arduino = None
        self.is_logging = False
        self.start_time = None
        self.duration_seconds = None
        self.data_filename = None
        self.active_channels: List[Tuple[int, str]] = []
        self.use_arduino_flag = False
        self.ambient_setpoint_value = None

        self.graph_window = None

        # For trend detection
        self.channel_history: Dict[int, List[float]] = {}
        self.trend_window = 10      # default number of recent samples
        self.trend_threshold = 3.0  # default °C range to call "stable"

        # status label handle for color control
        self.status_label = None

        self._build_vars()
        self._build_ui()
        self.set_status("Idle.")
        self.protocol("WM_DELETE_WINDOW", self.on_close)

    def _build_vars(self):
        self.test_name_var = tk.StringVar()
        self.tester_var = tk.StringVar()
        self.fixture_var = tk.StringVar()
        self.notes_var = tk.StringVar()

        self.include_cj_var = tk.BooleanVar(value=False)
        self.num_inputs_var = tk.IntVar(value=2)
        self.ch_name_vars = [tk.StringVar(value=f"CH{i}") for i in range(1, 9)]

        self.use_arduino_var = tk.BooleanVar(value=False)
        self.arduino_port_var = tk.StringVar(value="COM5")
        self.ambient_setpoint_var = tk.StringVar(value="25")

        today_str = datetime.now().strftime("%Y-%m-%d")
        default_name = f"{today_str} Thermal Test"
        self.base_name_var = tk.StringVar(value=default_name)
        self.duration_minutes_var = tk.StringVar(value="")

        self.status_var = tk.StringVar(value="Idle.")
        self.last_line_var = tk.StringVar(value="No data yet.")
        self.summary_var = tk.StringVar(value="No configuration yet.")
        self.summary_header_text = ""

        # Trends text (left, under channel inputs)
        self.channel_trends_var = tk.StringVar(
            value="Channel temperature trends will appear here once data arrives."
        )

        # Trend settings, adjustable in UI
        self.trend_window_var = tk.StringVar(value=str(self.trend_window))
        self.trend_threshold_var = tk.StringVar(value=f"{self.trend_threshold:.1f}")

        # Output naming / path
        self.append_datetime_var = tk.BooleanVar(value=False)
        self.output_path_var = tk.StringVar(value="")

    # status helper
    def set_status(self, text: str, is_error: bool = False):
        self.status_var.set(text)
        if self.status_label is not None:
            self.status_label.configure(foreground=("red" if is_error else "black"))

    def _build_ui(self):
        top = ttk.Frame(self, padding=10)
        top.pack(fill="x")
        ttk.Label(top, text="Thermal Temp Controller Logger",
                  font=("Century Gothic", 16, "bold")).pack(side="left")
        right_info = ttk.Frame(top)
        right_info.pack(side="right", anchor="e")
        ttk.Label(right_info, text="LUX Dynamics",
                  font=("Century Gothic", 12, "bold")).pack(anchor="e")
        ttk.Label(right_info, text="Kailani Puava Alarcon",
                  font=("Century Gothic", 10)).pack(anchor="e")

        main = ttk.Frame(self, padding=10)
        main.pack(fill="both", expand=True)

        # Left column
        left = ttk.Frame(main)
        left.pack(side="left", fill="y", padx=(0, 10))

        meta = ttk.LabelFrame(left, text="Test Metadata", padding=10)
        meta.pack(fill="x", pady=(0, 10))

        ttk.Label(meta, text="Test name:").grid(row=0, column=0, sticky="e")
        ttk.Entry(meta, textvariable=self.test_name_var, width=30).grid(row=0, column=1, sticky="w")

        ttk.Label(meta, text="Tester:").grid(row=1, column=0, sticky="e")
        ttk.Entry(meta, textvariable=self.tester_var, width=30).grid(row=1, column=1, sticky="w")

        ttk.Label(meta, text="Fixture:").grid(row=2, column=0, sticky="e")
        ttk.Entry(meta, textvariable=self.fixture_var, width=30).grid(row=2, column=1, sticky="w")

        ttk.Label(meta, text="Notes:").grid(row=3, column=0, sticky="ne")
        ttk.Entry(meta, textvariable=self.notes_var, width=30).grid(row=3, column=1, sticky="w")

        ch_frame = ttk.LabelFrame(left, text="TC-08 Channels", padding=10)
        ch_frame.pack(fill="x")

        ttk.Checkbutton(
            ch_frame,
            text="Include internal sensor (channel 0 / CJ)",
            variable=self.include_cj_var
        ).grid(row=0, column=0, columnspan=2, sticky="w")

        ttk.Label(ch_frame, text="# of inputs to log (1–8):").grid(row=1, column=0, sticky="e")
        ttk.Spinbox(ch_frame, from_=0, to=8, textvariable=self.num_inputs_var,
                    width=5).grid(row=1, column=1, sticky="w")

        row = 2
        for i in range(1, 9):
            ttk.Label(ch_frame, text=f"Input {i} name:").grid(row=row, column=0, sticky="e")
            ttk.Entry(ch_frame, textvariable=self.ch_name_vars[i - 1],
                      width=20).grid(row=row, column=1, sticky="w")
            row += 1

        # Trend settings UI
        ttk.Label(ch_frame, text="Trend window (samples):").grid(row=row, column=0, sticky="e")
        ttk.Entry(ch_frame, textvariable=self.trend_window_var, width=8).grid(
            row=row, column=1, sticky="w"
        )
        row += 1

        ttk.Label(ch_frame, text="Stable band (°C):").grid(row=row, column=0, sticky="e")
        ttk.Entry(ch_frame, textvariable=self.trend_threshold_var, width=8).grid(
            row=row, column=1, sticky="w"
        )
        row += 1

        # Trend text within channel section
        ttk.Label(
            ch_frame,
            textvariable=self.channel_trends_var,
            justify="left",
            foreground="gray"
        ).grid(row=row, column=0, columnspan=2, sticky="w", pady=(6, 0))

        # Right column
        right = ttk.Frame(main)
        right.pack(side="left", fill="both", expand=True)

        ar_frame = ttk.LabelFrame(right, text="Arduino Ambient Control", padding=10)
        ar_frame.pack(fill="x")

        ttk.Checkbutton(
            ar_frame, text="Use Arduino for ambient control/logging",
            variable=self.use_arduino_var
        ).grid(row=0, column=0, columnspan=2, sticky="w")

        ttk.Label(ar_frame, text="COM port (e.g. COM5 or 5):").grid(row=1, column=0, sticky="e")
        ttk.Entry(ar_frame, textvariable=self.arduino_port_var, width=10).grid(
            row=1, column=1, sticky="w"
        )

        ttk.Label(ar_frame, text="Ambient setpoint (°C):").grid(row=2, column=0, sticky="e")
        ttk.Entry(ar_frame, textvariable=self.ambient_setpoint_var, width=10).grid(
            row=2, column=1, sticky="w"
        )

        run_frame = ttk.LabelFrame(right, text="Run Settings", padding=10)
        run_frame.pack(fill="x", pady=(10, 0))

        ttk.Label(run_frame, text="Output folder:").grid(row=0, column=0, sticky="ne")
        self.output_folder_label = ttk.Label(
            run_frame,
            text=resolve_output_folder(),
            wraplength=350,
            justify="left"
        )
        self.output_folder_label.grid(row=0, column=1, sticky="w")

        ttk.Label(run_frame, text="Base file name:").grid(row=1, column=0, sticky="e")
        ttk.Entry(run_frame, textvariable=self.base_name_var, width=30).grid(
            row=1, column=1, sticky="w"
        )

        self.append_datetime_check = ttk.Checkbutton(
            run_frame,
            text="Append start time to file name",
            variable=self.append_datetime_var
        )
        self.append_datetime_check.grid(row=2, column=0, columnspan=2, sticky="w", pady=(4, 4))

        ttk.Label(run_frame, text="Duration (minutes, blank = unlimited):").grid(
            row=3, column=0, sticky="e"
        )
        ttk.Entry(run_frame, textvariable=self.duration_minutes_var, width=10).grid(
            row=3, column=1, sticky="w"
        )

        btn_frame = ttk.Frame(run_frame)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=(10, 0))

        self.start_button = ttk.Button(btn_frame, text="Start Logging", command=self.start_logging)
        self.start_button.pack(side="left", padx=(0, 10))

        self.stop_button = ttk.Button(btn_frame, text="Stop Logging", command=self.on_stop)
        self.stop_button.pack(side="left")
        self.stop_button["state"] = "disabled"

        # button to open graph window manually if you want
        open_graph_btn = ttk.Button(run_frame, text="Open Live Graph Window",
                                    command=self.ensure_graph_window)
        open_graph_btn.grid(row=5, column=0, columnspan=2, pady=(10, 0))

        ttk.Label(run_frame, text="Full output path:").grid(row=6, column=0, sticky="ne", pady=(8, 0))
        self.output_path_entry = ttk.Entry(run_frame, textvariable=self.output_path_var, width=40)
        self.output_path_entry.grid(row=6, column=1, sticky="w", pady=(8, 0))
        self.output_path_entry.configure(state="readonly")

        summary_frame = ttk.LabelFrame(right, text="Current Configuration", padding=10)
        summary_frame.pack(fill="both", expand=True, pady=(10, 0))
        ttk.Label(summary_frame, textvariable=self.summary_var,
                  justify="left", wraplength=400).pack(anchor="w")

        status_frame = ttk.LabelFrame(self, text="Status", padding=10)
        status_frame.pack(fill="x", side="bottom")

        self.status_label = ttk.Label(status_frame, textvariable=self.status_var)
        self.status_label.pack(anchor="w")
        ttk.Label(status_frame, text="Last reading:").pack(anchor="w")
        ttk.Label(status_frame, textvariable=self.last_line_var,
                  wraplength=800).pack(anchor="w")

    # --------- Logging control ---------- #

    def ensure_graph_window(self):
        if self.graph_window is None or not self.graph_window.winfo_exists():
            self.graph_window = LiveGraphWindow(self)
            if self.active_channels:
                self.graph_window.set_channels(self.active_channels)

    def start_logging(self):
        if self.is_logging:
            messagebox.showinfo("Logging", "Already logging.")
            return

        test_name = self.test_name_var.get().strip() or "Untitled Test"
        tester = self.tester_var.get().strip() or "Unknown"
        fixture = self.fixture_var.get().strip() or "N/A"
        notes = self.notes_var.get().strip()

        try:
            num_inputs = int(self.num_inputs_var.get())
        except ValueError:
            messagebox.showerror("Error", "Number of inputs must be a number between 0 and 8.")
            return
        if not (0 <= num_inputs <= 8):
            messagebox.showerror("Error", "Number of inputs must be between 0 and 8.")
            return

        channels: List[Tuple[int, str]] = []
        if self.include_cj_var.get():
            channels.append((0, "CJ"))

        for i in range(1, num_inputs + 1):
            name = self.ch_name_vars[i - 1].get().strip()
            if not name:
                name = f"CH{i}"
            channels.append((i, name))

        if not channels:
            messagebox.showerror("Error", "You must log at least one channel.")
            return

        self.active_channels = channels

        # read trend settings from UI
        try:
            tw_str = self.trend_window_var.get().strip()
            tw = int(tw_str)
            if tw < 2:
                raise ValueError
            self.trend_window = tw
        except Exception:
            messagebox.showerror(
                "Trend settings error",
                "Trend window (samples) must be an integer ≥ 2."
            )
            return

        try:
            band_str = self.trend_threshold_var.get().strip()
            band = float(band_str)
            if band <= 0:
                raise ValueError
            self.trend_threshold = band
        except Exception:
            messagebox.showerror(
                "Trend settings error",
                "Stable band (°C) must be a positive number."
            )
            return

        # Arduino
        self.use_arduino_flag = False
        self.ambient_setpoint_value = None
        if self.use_arduino_var.get():
            if not HAVE_SERIAL:
                messagebox.showerror(
                    "Arduino error",
                    "pyserial is not installed; cannot use Arduino.\nInstall it or uncheck 'Use Arduino'. Or get Kailani."
                )
                return

            port_input = self.arduino_port_var.get().strip()
            if not port_input:
                messagebox.showerror("Arduino error", "Please enter a COM port (e.g. COM5 or 5).")
                return

            if port_input.upper().startswith("COM"):
                port_name = port_input.upper()
            else:
                port_name = f"COM{port_input}"

            sp_str = self.ambient_setpoint_var.get().strip()
            try:
                sp = float(sp_str)
            except ValueError:
                messagebox.showerror("Arduino error. Get Kailani.", "Ambient setpoint must be a number.")
                return

            try:
                self.arduino = ArduinoInterface(port_name)
                self.use_arduino_flag = True
                self.ambient_setpoint_value = sp
                self.arduino.set_hold(sp)
            except Exception as e:
                messagebox.showerror(
                    "Arduino error. Get Kailani.",
                    f"Failed to connect to Arduino on {port_name}:\n{e}"
                )
                self.arduino = None
                self.use_arduino_flag = False

        # Output folder & filename
        output_folder = resolve_output_folder()
        self.output_folder_label.config(text=output_folder)

        base_name = self.base_name_var.get().strip()
        if not base_name:
            today_str = datetime.now().strftime("%Y-%m-%d")
            base_name = f"{today_str} Thermal Test"

        if self.append_datetime_var.get():
            time_str = datetime.now().strftime("%H-%M-%S")
            base_name = f"{base_name} {time_str}"

        self.base_name_var.set(base_name)

        self.data_filename = get_unique_csv_path(output_folder, base_name)
        self.output_path_var.set(self.data_filename)

        # Duration
        duration_str = self.duration_minutes_var.get().strip()
        if duration_str == "":
            self.duration_seconds = None
        else:
            try:
                minutes = float(duration_str)
                if minutes <= 0:
                    raise ValueError
                self.duration_seconds = minutes * 60.0
            except ValueError:
                messagebox.showerror(
                    "Error",
                    "Duration must be a positive number of minutes or left blank."
                )
                return

        # Open TC-08
        try:
            self.logger = TC08Interface()
        except Exception as e:
            messagebox.showerror("TC-08 error. Get Kailani.", f"Could not open TC-08:\n{e}")
            self.logger = None
            self.set_status("TC-08 error: could not open device.", is_error=True)
            return

        # Open CSV
        try:
            self.csv_file = open(self.data_filename, mode="w", newline="")
            self.csv_writer = csv.writer(self.csv_file)
        except Exception as e:
            messagebox.showerror("File error. Get Kailani.", f"Could not open CSV file for writing:\n{e}")
            if self.logger is not None:
                self.logger.close()
            self.logger = None
            self.set_status("File error: could not open CSV for writing.", is_error=True)
            return

        # Write header
        meta_text = (
            f"Test: {test_name} | "
            f"Tester: {tester} | "
            f"Fixture: {fixture} | "
            f"Notes: {notes}"
        )
        if self.ambient_setpoint_value is not None:
            meta_text += f" | Ambient setpoint: {self.ambient_setpoint_value:.2f} °C"
        self.csv_writer.writerow([meta_text])
        self.csv_writer.writerow([])

        header = ["timestamp"]
        if self.use_arduino_flag:
            header.append("Arduino_Temp")
        for _, name in self.active_channels:
            header.append(f"{name}_C")
        self.csv_writer.writerow(header)
        self.csv_file.flush()

        summary_lines = [
            f"Output file: {os.path.basename(self.data_filename)}",
            f"Test: {test_name}",
            f"Tester: {tester}",
            f"Fixture: {fixture}",
            (
                f"Ambient setpoint: {self.ambient_setpoint_value:.2f} °C"
                if self.ambient_setpoint_value is not None
                else "Ambient setpoint: N/A"
            ),
            "Channels:",
        ]
        for ch, name in self.active_channels:
            summary_lines.append(f"  Input {ch}: {name}")

        self.summary_header_text = "\n".join(summary_lines)
        self.summary_var.set(self.summary_header_text)

        # reset channel history each run
        self.channel_history = {}
        self.channel_trends_var.set(
            f"Channel temperature trends (last ~{self.trend_window} readings, "
            f"stable within ±{self.trend_threshold:.1f} °C) will appear here once data arrives."
        )

        # Set up graph window
        self.ensure_graph_window()
        if self.graph_window is not None and self.graph_window.winfo_exists():
            self.graph_window.set_channels(self.active_channels)

        self.start_time = time.time()
        self.is_logging = True
        self.set_status("Logging...")
        self.last_line_var.set("No data yet.")
        self.start_button["state"] = "disabled"
        self.stop_button["state"] = "normal"

        self.after(1000, self.poll_once)

    def poll_once(self):
        if not self.is_logging:
            return

        try:
            temps = self.logger.read() if self.logger is not None else {}
        except Exception as e:
            # Non-fatal: show in red status bar, keep logging and try again next tick
            self.set_status(f"TC-08 read error: {e}", is_error=True)
            self.after(1000, self.poll_once)
            return

        # if we get here, last read was OK → clear error status if previously set
        if self.status_var.get().startswith("TC-08 read error"):
            self.set_status("Logging...")

        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row = [ts]
        display_vals: List[str] = []

        if self.use_arduino_flag and self.arduino is not None:
            ar_temp, ar_hold, ar_pwm = self.arduino.poll()
            row.append(fmt_val(ar_temp))
            if ar_temp is not None:
                display_vals.append(
                    f"Arduino={ar_temp:.2f}°C (hold={ar_hold:.2f}°C, PWM={ar_pwm:.0f})"
                )
            else:
                display_vals.append("Arduino=NaN")

        for ch, name in self.active_channels:
            val = temps.get(ch, float("nan"))
            row.append(fmt_val(val))
            try:
                display_vals.append(f"{name}={val:.2f}°C")
            except TypeError:
                display_vals.append(f"{name}=NaN")

        if self.csv_writer is not None:
            try:
                self.csv_writer.writerow(row)
                self.csv_file.flush()
            except Exception as e:
                # writing error is serious: stop and show popup
                messagebox.showerror("File error. Get Kailani.", f"Error writing to CSV:\n{e}")
                self.set_status("File error while writing CSV.", is_error=True)
                self.stop_logging(error=True)
                return

        self.last_line_var.set(ts + " | " + "  ".join(display_vals))

        # update trends in channel section
        self.update_channel_trends(temps)

        if self.start_time is not None:
            elapsed = time.time() - self.start_time
        else:
            elapsed = 0.0

        if self.graph_window is not None and self.graph_window.winfo_exists():
            self.graph_window.add_sample(elapsed, temps)
        else:
            self.graph_window = None

        if self.duration_seconds is not None and self.start_time is not None:
            if elapsed >= self.duration_seconds:
                self.stop_logging(error=False)
                return

        self.after(1000, self.poll_once)

    # --------- Trend detection ------- ---------- #

    def update_channel_trends(self, temps: Dict[int, float]):
        """
        Decide if each channel is increasing / decreasing / stable
        (within self.trend_threshold °C) from last self.trend_window readings,
        and show this under TC-08 Channels on the left.
        """
        if not self.active_channels:
            return

        lines = [
            f"Channel temperature trends (last ~{self.trend_window} readings, "
            f"stable within ±{self.trend_threshold:.1f} °C):"
        ]

        for ch, name in self.active_channels:
            hlist = self.channel_history.setdefault(ch, [])

            val = temps.get(ch, None)
            # try to parse numeric
            try:
                v = float(val)
                if math.isnan(v):
                    raise ValueError
                hlist.append(v)
                if len(hlist) > self.trend_window:
                    del hlist[:-self.trend_window]
            except Exception:
                if not hlist:
                    lines.append(f"  {name}: no data")
                    continue

            if len(hlist) < 2:
                lines.append(f"  {name}: no data")
                continue

            vmin = min(hlist)
            vmax = max(hlist)
            if (vmax - vmin) <= self.trend_threshold:
                trend = "stable"
            else:
                delta = hlist[-1] - hlist[0]
                if delta > 0:
                    trend = "increasing"
                elif delta < 0:
                    trend = "decreasing"
                else:
                    trend = "stable"

            lines.append(f"  {name}: {trend}")

        self.channel_trends_var.set("\n".join(lines))

    # --------- Shutdown ---------- #

    def stop_logging(self, error: bool = False):
        if not self.is_logging:
            return

        self.is_logging = False
        self.start_button["state"] = "normal"
        self.stop_button["state"] = "disabled"

        try:
            if self.logger is not None:
                self.logger.close()
        except Exception:
            pass
        self.logger = None

        try:
            if self.csv_file is not None:
                self.csv_file.close()
        except Exception:
            pass
        self.csv_file = None
        self.csv_writer = None

        if not error and self.data_filename and HAVE_OPENPYXL:
            create_colored_excel(self.data_filename)

        if self.arduino is not None:
            try:
                self.arduino.close()
            except Exception:
                pass
            self.arduino = None

        self.set_status("Idle.")
        if not error and self.data_filename:
            messagebox.showinfo("Logging finished", f"Data saved to:\n{self.data_filename}")

    def on_stop(self):
        if self.is_logging:
            self.stop_logging(error=False)

    def on_close(self):
        if self.is_logging:
            if not messagebox.askyesno(
                "Quit",
                "Logging is still running. Stop and exit?"
            ):
                return
            self.stop_logging(error=True)

        if self.graph_window is not None and self.graph_window.winfo_exists():
            try:
                self.graph_window.destroy()
            except Exception:
                pass
            self.graph_window = None

        self.destroy()


if __name__ == "__main__":
    import traceback
    try:
        app = ThermalLoggerApp()
        app.mainloop()
    except Exception:
        traceback.print_exc()
        input("Error occurred, press Enter to exit...")
