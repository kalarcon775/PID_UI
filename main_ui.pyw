# main_ui.pyw
"""
LUX Dynamics Thermal Temp Controller Logger

- GUI for configuring tests, channels, Arduino ambient control.
- Logs TC-08 + Arduino data to CSV and colored Excel.
- Opens a separate live graph window (graph_window.LiveGraphWindow).
- Shows channel trends with direction, average, and delta, e.g.:
    CH1: stable (avg 54.2 °C, Δ=+0.7 °C)
"""

import sys
import subprocess
import time
import csv
import os
import math
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from typing import Dict, List, Tuple

from logger_core import (
    TC08Interface,
    ArduinoInterface,
    TREND_WINDOW_DEFAULT,
    TREND_THRESHOLD_DEFAULT,
    SAMPLE_INTERVAL,
)
from graph_window import LiveGraphWindow

DEPENDENCIES = [
    ("openpyxl", "openpyxl"),
    ("pywin32", "win32com.client"),
    ("pyserial", "serial"),
]

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Border, Side
    HAVE_OPENPYXL = True
except ImportError:
    Workbook = None
    PatternFill = None
    Border = None
    Side = None
    HAVE_OPENPYXL = False


def check_and_install_dependencies():
    missing: list[tuple[str, str]] = []

    for pip_name, import_name in DEPENDENCIES:
        try:
            __import__(import_name.split(".")[0])
        except ImportError:
            missing.append((pip_name, import_name))

    if not missing:
        return

    lines = [
        "This app is missing some Python packages needed for certain features:",
        "",
    ]
    for pip_name, import_name in missing:
        lines.append(f"• {pip_name}  (import '{import_name}')")
    lines.append("")
    lines.append("Would you like to try installing them automatically now?")

    msg = "\n".join(lines)

    if not messagebox.askyesno("Missing Python packages", msg):
        return

    python_exe = sys.executable or "python"
    failed: list[str] = []

    for pip_name, import_name in missing:
        try:
            subprocess.check_call([python_exe, "-m", "pip", "install", pip_name])
        except Exception:
            failed.append(pip_name)

    global HAVE_OPENPYXL, Workbook, PatternFill, Border, Side
    try:
        from openpyxl import Workbook as WB
        from openpyxl.styles import PatternFill as PF, Border as BD, Side as SD
        Workbook = WB
        PatternFill = PF
        Border = BD
        Side = SD
        HAVE_OPENPYXL = True
    except Exception:
        pass

    if failed:
        messagebox.showwarning(
            "Some packages not installed",
            "The app tried to install these packages but failed:\n\n"
            + "\n".join(f"• {name}" for name in failed)
            + "\n\nYou can still use most functionality, but some features may be disabled."
        )


APP_NAME = "LUX Thermal Logger"
SHORTCUT_NAME = "LUX Thermal Logger.lnk"
ICON_FILENAME = "lux_logo.ico"

COLOR_HIGHLIGHT = "#2aa84a"
COLOR_TEXT = "#ffffff"
COLOR_PANEL_TEXT = "#000000"
COLOR_MAIN_BG = "#000000"
COLOR_SECONDARY_BG = "#d9d9d9"
COLOR_FIELD_BG = "#ffffff"
COLOR_TAB_BG = "#777777"
COLOR_ERROR = "#ff6b6b"


def _get_windows_desktop() -> str:
    if os.name != "nt":
        return os.path.join(os.path.expanduser("~"), "Desktop")

    try:
        import ctypes
        from ctypes import wintypes

        SHGFP_TYPE_CURRENT = 0
        CSIDL_DESKTOPDIRECTORY = 0x10

        buf = ctypes.create_unicode_buffer(wintypes.MAX_PATH)
        ctypes.windll.shell32.SHGetFolderPathW(
            None, CSIDL_DESKTOPDIRECTORY, None, SHGFP_TYPE_CURRENT, buf
        )
        return buf.value
    except Exception:
        return os.path.join(os.path.expanduser("~"), "Desktop")


def ensure_desktop_shortcut():
    if os.name != "nt":
        return

    try:
        from win32com.client import Dispatch
    except ImportError:
        return

    desktop = _get_windows_desktop()
    shortcut_path = os.path.join(desktop, SHORTCUT_NAME)

    if os.path.exists(shortcut_path):
        return

    if getattr(sys, "frozen", False):
        target = sys.executable
        icon_location = f"{target},0"
    else:
        target = os.path.abspath(__file__)
        workdir = os.path.dirname(target)
        icon_path = os.path.join(workdir, ICON_FILENAME)
        if os.path.exists(icon_path):
            icon_location = f"{icon_path},0"
        else:
            icon_location = f"{target},0"

    workdir = os.path.dirname(target)

    shell = Dispatch("WScript.Shell")
    shortcut = shell.CreateShortCut(shortcut_path)
    shortcut.Targetpath = target
    shortcut.WorkingDirectory = workdir
    shortcut.IconLocation = icon_location
    shortcut.save()


OUTPUT_FOLDER = r"Z:\ENGINEERING\Product Development\Thermal Testing 2026"


def get_unique_csv_path(folder: str, base_name: str) -> str:
    path = os.path.join(folder, base_name + ".csv")
    if not os.path.exists(path):
        return path

    i = 1
    while True:
        alt = os.path.join(folder, f"{base_name}_{i}.csv")
        if not os.path.exists(alt):
            return alt
        i += 1


def resolve_output_folder() -> str:
    if os.path.isdir(OUTPUT_FOLDER):
        return OUTPUT_FOLDER

    fallback = os.path.join(os.getcwd(), "logs")
    os.makedirs(fallback, exist_ok=True)
    return fallback


def apply_column_colors(ws):
    if not HAVE_OPENPYXL or PatternFill is None:
        return

    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value == "Data_Sample":
                header_row_idx = cell.row
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        return

    header_cells = list(ws[header_row_idx])
    num_cols = len(header_cells)

    if num_cols == 0:
        return

    palette = [
        "FFCCCC", "FFE5CC", "FFF2CC", "E5FFCC",
        "CCFFFF", "CCE5FF", "E5CCFF", "FFCCF2", "E6E6FA",
    ]

    bold_border = Border(
        left=Side(style="medium", color="000000"),
        right=Side(style="medium", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col_idx, cell in enumerate(header_cells, start=1):
        color_hex = palette[(col_idx - 1) % len(palette)]
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

        for row_idx in range(header_row_idx, ws.max_row + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.fill = fill
            c.border = bold_border


def create_colored_excel(csv_path: str):
    if not HAVE_OPENPYXL or Workbook is None:
        print("openpyxl not available. Skipping colored Excel export.")
        return

    xlsx_path = os.path.splitext(csv_path)[0] + ".xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "TC08 Log"

    with open(csv_path, newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    apply_column_colors(ws)
    wb.save(xlsx_path)
    print(f"Colored Excel copy saved as:\n  {xlsx_path}")


def fmt_val(val):
    try:
        if val is None:
            return ""
        if isinstance(val, float) and math.isnan(val):
            return ""
        return f"{float(val):.2f}"
    except (TypeError, ValueError):
        return ""


class ThermalLoggerApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("LUX Thermal Logger")
        self.geometry("1040x720")

        self.logger = None
        self.csv_file = None
        self.csv_writer = None
        self.arduino = None

        self.is_logging = False
        self.start_time = None
        self.sample_count = 0
        self.duration_seconds = None
        self.data_filename = None
        self.active_channels: List[Tuple[int, str]] = []
        self.use_arduino_flag = False
        self.ambient_setpoint_value = None

        self.ambient_feedback_channel = None
        self.step_mode_enabled = False
        self.step_setpoints = []
        self.step_hold_seconds = None
        self.last_sent_setpoint = None

        self.graph_window: LiveGraphWindow | None = None

        self.channel_history: Dict[int, List[float]] = {}
        self.trend_window = TREND_WINDOW_DEFAULT
        self.trend_threshold = TREND_THRESHOLD_DEFAULT

        self.status_label = None
        self.summary_header_text = ""

        self._build_vars()
        self._configure_theme()
        self._build_ui()
        self.set_status("Idle.")
        self.after(0, self._post_init)
        self.protocol("WM_DELETE_WINDOW", self.on_close)

    def _post_init(self):
        try:
            check_and_install_dependencies()
        except Exception:
            pass

        try:
            ensure_desktop_shortcut()
        except Exception:
            pass

    def _build_vars(self):
        self.test_name_var = tk.StringVar()
        self.tester_var = tk.StringVar()
        self.fixture_var = tk.StringVar()
        self.notes_var = tk.StringVar()

        self.include_cj_var = tk.BooleanVar(value=False)
        self.num_inputs_var = tk.IntVar(value=2)
        self.ch_name_vars = [tk.StringVar(value=f"CH{i}") for i in range(1, 9)]

        self.use_arduino_var = tk.BooleanVar(value=False)
        self.arduino_port_var = tk.StringVar(value="COM5")
        self.ambient_setpoint_var = tk.StringVar(value="25")
        self.ambient_feedback_channel_var = tk.IntVar(value=1)

        self.step_mode_var = tk.BooleanVar(value=False)
        self.step_start_var = tk.StringVar(value="40")
        self.step_stop_var = tk.StringVar(value="60")
        self.step_size_var = tk.StringVar(value="5")
        self.step_hold_hours_var = tk.StringVar(value="3")

        today_str = datetime.now().strftime("%Y-%m-%d")
        default_name = f"{today_str} Thermal Test"
        self.base_name_var = tk.StringVar(value=default_name)
        self.duration_minutes_var = tk.StringVar(value="")

        self.status_var = tk.StringVar(value="Idle.")
        self.last_line_var = tk.StringVar(value="No data yet.")
        self.summary_var = tk.StringVar(value="No configuration yet.")

        self.channel_trends_var = tk.StringVar(
            value="Channel temperature trends will appear here once data arrives."
        )

        self.trend_window_var = tk.StringVar(value=str(self.trend_window))
        self.trend_threshold_var = tk.StringVar(value=f"{self.trend_threshold:.1f}")

        self.append_datetime_var = tk.BooleanVar(value=False)
        self.output_path_var = tk.StringVar(value="")

    def set_status(self, text: str, is_error: bool = False):
        self.status_var.set(text)
        if self.status_label is not None:
            self.status_label.configure(foreground=(COLOR_ERROR if is_error else COLOR_TEXT))

    def _configure_theme(self):
        self.configure(bg=COLOR_MAIN_BG)

        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        base_font = ("Segoe UI", 10)
        header_font = ("Century Gothic", 16, "bold")
        subheader_font = ("Century Gothic", 12, "bold")

        style.configure(".", font=base_font)
        style.configure("TFrame", background=COLOR_MAIN_BG)
        style.configure("Secondary.TFrame", background=COLOR_SECONDARY_BG)
        style.configure("TLabel", background=COLOR_SECONDARY_BG, foreground=COLOR_PANEL_TEXT)
        style.configure("Main.TLabel", background=COLOR_MAIN_BG, foreground=COLOR_TEXT)
        style.configure("Secondary.TLabel", background=COLOR_SECONDARY_BG, foreground=COLOR_PANEL_TEXT)
        style.configure("Header.TLabel", background=COLOR_MAIN_BG, foreground=COLOR_TEXT, font=header_font)
        style.configure("Subheader.TLabel", background=COLOR_MAIN_BG, foreground=COLOR_HIGHLIGHT, font=subheader_font)
        style.configure("PanelHeader.TLabel", background=COLOR_SECONDARY_BG, foreground=COLOR_PANEL_TEXT, font=subheader_font)
        style.configure("Muted.TLabel", background=COLOR_SECONDARY_BG, foreground=COLOR_PANEL_TEXT)
        style.configure(
            "TLabelFrame",
            background=COLOR_SECONDARY_BG,
            foreground=COLOR_PANEL_TEXT,
            bordercolor=COLOR_HIGHLIGHT,
            relief="solid",
        )
        style.configure(
            "TLabelFrame.Label",
            background=COLOR_SECONDARY_BG,
            foreground=COLOR_PANEL_TEXT,
            font=("Segoe UI", 10, "bold"),
        )
        style.configure("TCheckbutton", background=COLOR_SECONDARY_BG, foreground=COLOR_PANEL_TEXT)
        style.map(
            "TCheckbutton",
            background=[("active", COLOR_SECONDARY_BG), ("selected", COLOR_SECONDARY_BG)],
            foreground=[("active", COLOR_PANEL_TEXT), ("selected", COLOR_PANEL_TEXT)],
        )
        style.configure(
            "TButton",
            background=COLOR_HIGHLIGHT,
            foreground=COLOR_TEXT,
            bordercolor=COLOR_HIGHLIGHT,
            focusthickness=2,
            focuscolor=COLOR_TEXT,
            padding=(10, 5),
        )
        style.map(
            "TButton",
            background=[("disabled", COLOR_TAB_BG), ("pressed", COLOR_TAB_BG), ("active", COLOR_HIGHLIGHT)],
            foreground=[("disabled", "#dddddd"), ("pressed", COLOR_TEXT), ("active", COLOR_TEXT)],
        )
        style.configure("TNotebook", background=COLOR_MAIN_BG, borderwidth=0)
        style.configure(
            "TNotebook.Tab",
            background=COLOR_TAB_BG,
            foreground=COLOR_TEXT,
            padding=(16, 8),
        )
        style.map(
            "TNotebook.Tab",
            background=[("selected", COLOR_HIGHLIGHT), ("active", COLOR_HIGHLIGHT)],
            foreground=[("selected", COLOR_TEXT), ("active", COLOR_TEXT)],
        )
        style.configure(
            "TEntry",
            fieldbackground=COLOR_FIELD_BG,
            foreground=COLOR_PANEL_TEXT,
            insertcolor=COLOR_PANEL_TEXT,
            bordercolor=COLOR_HIGHLIGHT,
            lightcolor=COLOR_HIGHLIGHT,
            darkcolor=COLOR_HIGHLIGHT,
        )
        style.map(
            "TEntry",
            fieldbackground=[("readonly", COLOR_FIELD_BG), ("disabled", COLOR_SECONDARY_BG)],
            foreground=[("readonly", COLOR_PANEL_TEXT), ("disabled", "#222222")],
        )
        style.configure(
            "TSpinbox",
            fieldbackground=COLOR_FIELD_BG,
            foreground=COLOR_PANEL_TEXT,
            arrowcolor=COLOR_PANEL_TEXT,
            bordercolor=COLOR_HIGHLIGHT,
        )

    def _build_ui(self):
        top = ttk.Frame(self, padding=10)
        top.pack(fill="x")

        ttk.Label(
            top,
            text="Thermal Temp Controller Logger",
            style="Header.TLabel"
        ).pack(side="left")

        right_info = ttk.Frame(top)
        right_info.pack(side="right", anchor="e")

        ttk.Label(
            right_info,
            text="LUX Dynamics",
            style="Subheader.TLabel"
        ).pack(anchor="e")

        ttk.Label(
            right_info,
            text="Kailani Puava Alarcon",
            style="Main.TLabel"
        ).pack(anchor="e")

        action_bar = ttk.Frame(self, padding=(10, 0, 10, 10))
        action_bar.pack(fill="x")

        self.start_button = ttk.Button(action_bar, text="Start Logging", command=self.start_logging)
        self.start_button.pack(side="left", padx=(0, 10))

        self.stop_button = ttk.Button(action_bar, text="Stop Logging", command=self.on_stop)
        self.stop_button.pack(side="left", padx=(0, 10))
        self.stop_button["state"] = "disabled"

        ttk.Button(
            action_bar,
            text="Open Live Graph Window",
            command=self.ensure_graph_window
        ).pack(side="left")

        self.status_label = ttk.Label(action_bar, textvariable=self.status_var, style="Main.TLabel")
        self.status_label.pack(side="right")

        main = ttk.Frame(self, padding=10)
        main.pack(fill="both", expand=True)

        notebook = ttk.Notebook(main)
        notebook.pack(fill="both", expand=True)

        test_tab = ttk.Frame(notebook, padding=12)
        channels_tab = ttk.Frame(notebook, padding=12)
        ambient_tab = ttk.Frame(notebook, padding=12)
        summary_tab = ttk.Frame(notebook, padding=12)

        notebook.add(test_tab, text="Test Info")
        notebook.add(channels_tab, text="TC-08 Channels")
        notebook.add(ambient_tab, text="Ambient Control")
        notebook.add(summary_tab, text="Summary")

        meta = ttk.LabelFrame(test_tab, text="Test Metadata", padding=10)
        meta.pack(fill="x", pady=(0, 10))
        meta.columnconfigure(1, weight=1)

        ttk.Label(meta, text="Test name:").grid(row=0, column=0, sticky="e")
        ttk.Entry(meta, textvariable=self.test_name_var, width=42).grid(row=0, column=1, sticky="ew", padx=(8, 0), pady=2)

        ttk.Label(meta, text="Tester:").grid(row=1, column=0, sticky="e")
        ttk.Entry(meta, textvariable=self.tester_var, width=42).grid(row=1, column=1, sticky="ew", padx=(8, 0), pady=2)

        ttk.Label(meta, text="Fixture:").grid(row=2, column=0, sticky="e")
        ttk.Entry(meta, textvariable=self.fixture_var, width=42).grid(row=2, column=1, sticky="ew", padx=(8, 0), pady=2)

        ttk.Label(meta, text="Notes:").grid(row=3, column=0, sticky="ne")
        ttk.Entry(meta, textvariable=self.notes_var, width=42).grid(row=3, column=1, sticky="ew", padx=(8, 0), pady=2)

        ch_frame = ttk.LabelFrame(channels_tab, text="TC-08 Channels + Trends", padding=10)
        ch_frame.pack(fill="x")
        ch_frame.columnconfigure(1, weight=1)

        ttk.Checkbutton(
            ch_frame,
            text="Include internal sensor (channel 0 / CJ)",
            variable=self.include_cj_var
        ).grid(row=0, column=0, columnspan=2, sticky="w")

        ttk.Label(ch_frame, text="# of inputs to log (1 to 8):").grid(row=1, column=0, sticky="e")

        ttk.Spinbox(
            ch_frame,
            from_=0,
            to=8,
            textvariable=self.num_inputs_var,
            width=5
        ).grid(row=1, column=1, sticky="w")

        row = 2
        for i in range(1, 9):
            ttk.Label(ch_frame, text=f"Input {i} name:").grid(row=row, column=0, sticky="e")
            ttk.Entry(ch_frame, textvariable=self.ch_name_vars[i - 1], width=20).grid(
                row=row,
                column=1,
                sticky="w"
            )
            row += 1

        ttk.Label(ch_frame, text="Trend window (samples):").grid(row=row, column=0, sticky="e")
        ttk.Entry(ch_frame, textvariable=self.trend_window_var, width=8).grid(
            row=row,
            column=1,
            sticky="w"
        )
        row += 1

        ttk.Label(ch_frame, text="Stable band (°C):").grid(row=row, column=0, sticky="e")
        ttk.Entry(ch_frame, textvariable=self.trend_threshold_var, width=8).grid(
            row=row,
            column=1,
            sticky="w"
        )
        row += 1

        ttk.Label(
            ch_frame,
            textvariable=self.channel_trends_var,
            justify="left",
            style="Muted.TLabel"
        ).grid(row=row, column=0, columnspan=2, sticky="w", pady=(6, 0))

        ar_frame = ttk.LabelFrame(ambient_tab, text="Arduino Ambient Control", padding=10)
        ar_frame.pack(fill="x")
        ar_frame.columnconfigure(1, weight=1)

        ttk.Checkbutton(
            ar_frame,
            text="Use Arduino for ambient control/logging",
            variable=self.use_arduino_var,
            command=self._update_arduino_controls
        ).grid(row=0, column=0, columnspan=2, sticky="w")

        ttk.Label(ar_frame, text="COM port (e.g. COM5 or 5):").grid(row=1, column=0, sticky="e")

        ttk.Entry(
            ar_frame,
            textvariable=self.arduino_port_var,
            width=12
        ).grid(row=1, column=1, sticky="w")

        ttk.Label(ar_frame, text="Ambient setpoint (°C):").grid(row=2, column=0, sticky="e")

        ttk.Entry(
            ar_frame,
            textvariable=self.ambient_setpoint_var,
            width=12
        ).grid(row=2, column=1, sticky="w")

        ttk.Label(ar_frame, text="Ambient TC-08 input channel:").grid(row=3, column=0, sticky="e")

        ttk.Spinbox(
            ar_frame,
            from_=1,
            to=8,
            textvariable=self.ambient_feedback_channel_var,
            width=5
        ).grid(row=3, column=1, sticky="w")

        self.step_mode_check = ttk.Checkbutton(
            ar_frame,
            text="Use step schedule",
            variable=self.step_mode_var,
            command=self._update_arduino_controls
        )
        self.step_mode_check.grid(row=4, column=0, columnspan=2, sticky="w")

        ttk.Label(ar_frame, text="Start °C:").grid(row=5, column=0, sticky="e")
        self.step_start_entry = ttk.Entry(ar_frame, textvariable=self.step_start_var, width=8)
        self.step_start_entry.grid(row=5, column=1, sticky="w")

        ttk.Label(ar_frame, text="Stop °C:").grid(row=6, column=0, sticky="e")
        self.step_stop_entry = ttk.Entry(ar_frame, textvariable=self.step_stop_var, width=8)
        self.step_stop_entry.grid(row=6, column=1, sticky="w")

        ttk.Label(ar_frame, text="Step °C:").grid(row=7, column=0, sticky="e")
        self.step_size_entry = ttk.Entry(ar_frame, textvariable=self.step_size_var, width=8)
        self.step_size_entry.grid(row=7, column=1, sticky="w")

        self.step_hold_label = ttk.Label(ar_frame, text="Hold hours per step:")
        self.step_hold_label.grid(row=8, column=0, sticky="e")
        self.step_hold_entry = ttk.Entry(ar_frame, textvariable=self.step_hold_hours_var, width=8)
        self.step_hold_entry.grid(row=8, column=1, sticky="w")

        self.step_schedule_controls = [
            self.step_start_entry,
            self.step_stop_entry,
            self.step_size_entry,
            self.step_hold_entry,
        ]
        self._update_arduino_controls()

        run_frame = ttk.LabelFrame(test_tab, text="Run Settings", padding=10)
        run_frame.pack(fill="x", pady=(10, 0))
        run_frame.columnconfigure(1, weight=1)

        ttk.Label(run_frame, text="Output folder:").grid(row=0, column=0, sticky="ne")

        self.output_folder_label = ttk.Label(
            run_frame,
            text=resolve_output_folder(),
            wraplength=420,
            justify="left"
        )
        self.output_folder_label.grid(row=0, column=1, sticky="ew", padx=(8, 0), pady=2)

        ttk.Label(run_frame, text="Base file name:").grid(row=1, column=0, sticky="e")

        ttk.Entry(
            run_frame,
            textvariable=self.base_name_var,
            width=36
        ).grid(row=1, column=1, sticky="ew", padx=(8, 0), pady=2)

        self.append_datetime_check = ttk.Checkbutton(
            run_frame,
            text="Append start time to file name",
            variable=self.append_datetime_var
        )
        self.append_datetime_check.grid(row=2, column=0, columnspan=2, sticky="w", pady=(4, 4))

        ttk.Label(run_frame, text="Duration (minutes, blank = unlimited):").grid(
            row=3,
            column=0,
            sticky="e"
        )

        ttk.Entry(
            run_frame,
            textvariable=self.duration_minutes_var,
            width=12
        ).grid(row=3, column=1, sticky="w", padx=(8, 0), pady=2)

        ttk.Label(run_frame, text="Full output path:").grid(row=4, column=0, sticky="ne", pady=(8, 0))

        self.output_path_entry = ttk.Entry(
            run_frame,
            textvariable=self.output_path_var,
            width=48
        )
        self.output_path_entry.grid(row=4, column=1, sticky="ew", padx=(8, 0), pady=(8, 0))
        self.output_path_entry.configure(state="readonly")

        summary_frame = ttk.LabelFrame(summary_tab, text="Current Configuration", padding=10)
        summary_frame.pack(fill="both", expand=True, pady=(10, 0))

        ttk.Label(
            summary_frame,
            textvariable=self.summary_var,
            justify="left",
            wraplength=760
        ).pack(anchor="w")

        status_frame = ttk.LabelFrame(summary_tab, text="Status", padding=10)
        status_frame.pack(fill="x", pady=(10, 0))

        ttk.Label(status_frame, textvariable=self.status_var, style="PanelHeader.TLabel").pack(anchor="w")

        ttk.Label(status_frame, text="Last reading:").pack(anchor="w")

        ttk.Label(
            status_frame,
            textvariable=self.last_line_var,
            wraplength=900
        ).pack(anchor="w")

    def _update_arduino_controls(self):
        arduino_enabled = self.use_arduino_var.get()

        if not arduino_enabled:
            self.step_mode_var.set(False)

        step_check_state = "normal" if arduino_enabled else "disabled"
        if hasattr(self, "step_mode_check"):
            self.step_mode_check.configure(state=step_check_state)

        step_control_state = "normal" if arduino_enabled and self.step_mode_var.get() else "disabled"
        for widget in getattr(self, "step_schedule_controls", []):
            widget.configure(state=step_control_state)

    def ensure_graph_window(self):
        if self.graph_window is None or not self.graph_window.winfo_exists():
            self.graph_window = LiveGraphWindow(self)
            if self.active_channels:
                self.graph_window.set_channels(self.active_channels)

    def build_step_setpoints(self, start_c: float, stop_c: float, step_c: float):
        if step_c <= 0:
            raise ValueError("Step size must be positive.")

        points = []
        current = start_c

        if start_c <= stop_c:
            while current <= stop_c + 1e-9:
                points.append(round(current, 2))
                current += step_c
        else:
            while current >= stop_c - 1e-9:
                points.append(round(current, 2))
                current -= step_c

        if not points or points[-1] != round(stop_c, 2):
            points.append(round(stop_c, 2))

        return points

    def get_step_setpoint(self, elapsed_seconds: float):
        if not self.step_mode_enabled:
            return self.ambient_setpoint_value, False

        if not self.step_setpoints or not self.step_hold_seconds:
            return self.ambient_setpoint_value, False

        step_index = int(elapsed_seconds // self.step_hold_seconds)

        if step_index >= len(self.step_setpoints):
            return self.step_setpoints[-1], True

        return self.step_setpoints[step_index], False

    def start_logging(self):
        if self.is_logging:
            messagebox.showinfo("Logging", "Already logging.")
            return

        test_name = self.test_name_var.get().strip() or "Untitled Test"
        tester = self.tester_var.get().strip() or "Unknown"
        fixture = self.fixture_var.get().strip() or "N/A"
        notes = self.notes_var.get().strip()

        try:
            num_inputs = int(self.num_inputs_var.get())
        except ValueError:
            messagebox.showerror("Error", "Number of inputs must be a number between 0 and 8.")
            return

        if not (0 <= num_inputs <= 8):
            messagebox.showerror("Error", "Number of inputs must be between 0 and 8.")
            return

        channels: List[Tuple[int, str]] = []

        if self.include_cj_var.get():
            channels.append((0, "CJ"))

        for i in range(1, num_inputs + 1):
            name = self.ch_name_vars[i - 1].get().strip()
            if not name:
                name = f"CH{i}"
            channels.append((i, name))

        if not channels:
            messagebox.showerror("Error", "You must log at least one channel.")
            return

        self.active_channels = channels

        try:
            tw_str = self.trend_window_var.get().strip()
            tw = int(tw_str)
            if tw < 2:
                raise ValueError
            self.trend_window = tw
        except Exception:
            messagebox.showerror(
                "Trend settings error",
                "Trend window (samples) must be an integer of at least 2."
            )
            return

        try:
            band_str = self.trend_threshold_var.get().strip()
            band = float(band_str)
            if band <= 0:
                raise ValueError
            self.trend_threshold = band
        except Exception:
            messagebox.showerror(
                "Trend settings error",
                "Stable band (°C) must be a positive number."
            )
            return

        self.use_arduino_flag = False
        self.ambient_setpoint_value = None
        self.ambient_feedback_channel = None
        self.step_mode_enabled = False
        self.step_setpoints = []
        self.step_hold_seconds = None
        self.last_sent_setpoint = None

        if self.use_arduino_var.get():
            port_input = self.arduino_port_var.get().strip()

            if not port_input:
                messagebox.showerror("Arduino error", "Please enter a COM port, for example COM5 or 5.")
                return

            if port_input.upper().startswith("COM"):
                port_name = port_input.upper()
            else:
                port_name = f"COM{port_input}"

            try:
                feedback_ch = int(self.ambient_feedback_channel_var.get())
                if not (1 <= feedback_ch <= 8):
                    raise ValueError
                self.ambient_feedback_channel = feedback_ch
            except Exception:
                messagebox.showerror(
                    "Arduino error",
                    "Ambient feedback channel must be a TC-08 thermocouple input from 1 to 8."
                )
                return

            if self.step_mode_var.get():
                try:
                    start_c = float(self.step_start_var.get().strip())
                    stop_c = float(self.step_stop_var.get().strip())
                    step_c = float(self.step_size_var.get().strip())
                    hold_hours = float(self.step_hold_hours_var.get().strip())

                    if hold_hours <= 0:
                        raise ValueError

                    self.step_setpoints = self.build_step_setpoints(start_c, stop_c, step_c)
                    self.step_hold_seconds = hold_hours * 3600.0
                    self.step_mode_enabled = True
                    sp = self.step_setpoints[0]

                except Exception:
                    messagebox.showerror(
                        "Step schedule error",
                        "Use valid numbers. Hold time must be a positive number of hours."
                    )
                    return
            else:
                sp_str = self.ambient_setpoint_var.get().strip()
                try:
                    sp = float(sp_str)
                except ValueError:
                    messagebox.showerror("Arduino error. Get Kailani.", "Ambient setpoint must be a number.")
                    return

            try:
                self.arduino = ArduinoInterface(port_name)
                self.use_arduino_flag = True
                self.ambient_setpoint_value = sp
                self.arduino.set_hold(sp)
                self.last_sent_setpoint = sp
            except Exception as e:
                messagebox.showerror(
                    "Arduino error. Get Kailani.",
                    f"Failed to connect to Arduino on {port_name}:\n{e}"
                )
                self.arduino = None
                self.use_arduino_flag = False
                return

        output_folder = resolve_output_folder()
        self.output_folder_label.config(text=output_folder)

        base_name = self.base_name_var.get().strip()

        if not base_name:
            today_str = datetime.now().strftime("%Y-%m-%d")
            base_name = f"{today_str} Thermal Test"

        if self.append_datetime_var.get():
            time_str = datetime.now().strftime("%H-%M-%S")
            base_name = f"{base_name} {time_str}"

        self.base_name_var.set(base_name)
        self.data_filename = get_unique_csv_path(output_folder, base_name)
        self.output_path_var.set(self.data_filename)

        duration_str = self.duration_minutes_var.get().strip()

        if duration_str == "":
            self.duration_seconds = None
        else:
            try:
                minutes = float(duration_str)
                if minutes <= 0:
                    raise ValueError
                self.duration_seconds = minutes * 60.0
            except ValueError:
                messagebox.showerror(
                    "Error",
                    "Duration must be a positive number of minutes or left blank."
                )
                return

        try:
            self.logger = TC08Interface()
        except Exception as e:
            messagebox.showerror("TC-08 error. Get Kailani.", f"Could not open TC-08:\n{e}")
            self.logger = None
            self.set_status("TC-08 error: could not open device.", is_error=True)
            return

        try:
            self.csv_file = open(self.data_filename, mode="w", newline="")
            self.csv_writer = csv.writer(self.csv_file)
        except Exception as e:
            messagebox.showerror("File error. Get Kailani.", f"Could not open CSV file for writing:\n{e}")
            if self.logger is not None:
                self.logger.close()
            self.logger = None
            self.set_status("File error: could not open CSV for writing.", is_error=True)
            return

        meta_text = (
            f"Test: {test_name} | "
            f"Tester: {tester} | "
            f"Fixture: {fixture} | "
            f"Notes: {notes}"
        )

        if self.ambient_setpoint_value is not None:
            meta_text += f" | Ambient setpoint: {self.ambient_setpoint_value:.2f} °C"

        if self.use_arduino_flag and self.ambient_feedback_channel is not None:
            meta_text += f" | Ambient feedback: TC-08 CH{self.ambient_feedback_channel}"

        if self.step_mode_enabled:
            meta_text += (
                f" | Step schedule: {self.step_setpoints} °C, "
                f"hold {self.step_hold_seconds / 3600.0:.2f} hours each"
            )

        self.csv_writer.writerow([meta_text])
        self.csv_writer.writerow([])

        self.sample_count = 0

        header = ["Data_Sample"]

        if self.use_arduino_flag:
            header.extend(["Ambient_Feedback_C", "Ambient_Setpoint_C", "Arduino_PWM"])

        for _, name in self.active_channels:
            header.append(f"{name}_C")

        self.csv_writer.writerow(header)
        self.csv_file.flush()

        summary_lines = [
            f"Output file: {os.path.basename(self.data_filename)}",
            f"Test: {test_name}",
            f"Tester: {tester}",
            f"Fixture: {fixture}",
            (
                f"Ambient setpoint: {self.ambient_setpoint_value:.2f} °C"
                if self.ambient_setpoint_value is not None
                else "Ambient setpoint: N/A"
            ),
        ]

        if self.use_arduino_flag:
            summary_lines.append(f"Ambient feedback channel: TC-08 CH{self.ambient_feedback_channel}")

            if self.step_mode_enabled:
                summary_lines.append(f"Step schedule: {self.step_setpoints} °C")
                summary_lines.append(f"Hold time per step: {self.step_hold_seconds / 3600.0:.2f} hours")

        summary_lines.append("Channels:")

        for ch, name in self.active_channels:
            summary_lines.append(f"  Input {ch}: {name}")

        self.summary_header_text = "\n".join(summary_lines)
        self.summary_var.set(self.summary_header_text)

        self.channel_history = {}
        self.channel_trends_var.set(
            f"Channel temperature trends (last ~{self.trend_window} readings, "
            f"stable within ±{self.trend_threshold:.1f} °C) will appear here once data arrives."
        )

        self.ensure_graph_window()

        if self.graph_window is not None and self.graph_window.winfo_exists():
            self.graph_window.set_channels(self.active_channels)

        self.start_time = time.time()
        self.is_logging = True
        self.set_status("Logging...")
        self.last_line_var.set("No data yet.")
        self.start_button["state"] = "disabled"
        self.stop_button["state"] = "normal"

        self.after(int(SAMPLE_INTERVAL * 1000), self.poll_once)

    def poll_once(self):
        if not self.is_logging:
            return

        try:
            temps = self.logger.read() if self.logger is not None else {}
        except Exception as e:
            self.set_status(f"TC-08 read error: {e}", is_error=True)
            self.after(int(SAMPLE_INTERVAL * 1000), self.poll_once)
            return

        if self.status_var.get().startswith("TC-08 read error"):
            self.set_status("Logging...")

        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.sample_count += 1
        row = [self.sample_count]
        display_vals: List[str] = []

        if self.start_time is not None:
            elapsed = time.time() - self.start_time
        else:
            elapsed = 0.0

        step_complete = False

        if self.use_arduino_flag and self.arduino is not None:
            ambient_temp = temps.get(self.ambient_feedback_channel, float("nan"))

            try:
                ambient_temp = float(ambient_temp)
                if math.isnan(ambient_temp):
                    raise ValueError
            except Exception:
                self.set_status(
                    f"Ambient feedback error: TC-08 input {self.ambient_feedback_channel} has no valid temperature.",
                    is_error=True
                )
                ambient_temp = None

            if ambient_temp is not None and self.status_var.get().startswith("Ambient feedback error"):
                self.set_status("Logging...")

            active_setpoint, step_complete = self.get_step_setpoint(elapsed)

            if active_setpoint is not None:
                if self.last_sent_setpoint is None or abs(active_setpoint - self.last_sent_setpoint) >= 0.01:
                    self.arduino.set_hold(active_setpoint)
                    self.last_sent_setpoint = active_setpoint

            if ambient_temp is not None:
                self.arduino.send_ambient(ambient_temp)

            ar_ambient, ar_hold, ar_pwm, ar_status = self.arduino.poll()

            row.extend([
                fmt_val(ambient_temp),
                fmt_val(active_setpoint),
                fmt_val(ar_pwm),
            ])

            display_vals.append(
                f"Ambient TC08 CH{self.ambient_feedback_channel}={fmt_val(ambient_temp)}°C "
                f"(setpoint={fmt_val(active_setpoint)}°C, PWM={fmt_val(ar_pwm)})"
            )

            if ar_status:
                display_vals.append(f"Arduino status={ar_status}")

        for ch, name in self.active_channels:
            val = temps.get(ch, float("nan"))
            row.append(fmt_val(val))

            try:
                display_vals.append(f"{name}={val:.2f}°C")
            except TypeError:
                display_vals.append(f"{name}=NaN")

        if self.csv_writer is not None:
            try:
                self.csv_writer.writerow(row)
                self.csv_file.flush()
            except Exception as e:
                messagebox.showerror("File error. Get Kailani.", f"Error writing to CSV:\n{e}")
                self.set_status("File error while writing CSV.", is_error=True)
                self.stop_logging(error=True)
                return

        self.last_line_var.set(ts + " | " + "  ".join(display_vals))

        self.update_channel_trends(temps)

        if self.graph_window is not None and self.graph_window.winfo_exists():
            self.graph_window.add_sample(elapsed, temps)
        else:
            self.graph_window = None

        if step_complete:
            self.stop_logging(error=False)
            return

        if self.duration_seconds is not None and self.start_time is not None:
            if elapsed >= self.duration_seconds:
                self.stop_logging(error=False)
                return

        self.after(int(SAMPLE_INTERVAL * 1000), self.poll_once)

    def update_channel_trends(self, temps: Dict[int, float]):
        if not self.active_channels:
            return

        lines = [
            f"Channel temperature trends (last ~{self.trend_window} readings, "
            f"stable within ±{self.trend_threshold:.1f} °C):"
        ]

        for ch, name in self.active_channels:
            hlist = self.channel_history.setdefault(ch, [])
            val = temps.get(ch, None)

            try:
                v = float(val)
                if math.isnan(v):
                    raise ValueError
                hlist.append(v)

                if len(hlist) > self.trend_window:
                    del hlist[:-self.trend_window]

            except Exception:
                if not hlist:
                    lines.append(f"  {name}: no data")
                    continue

            if len(hlist) < 2:
                lines.append(f"  {name}: no data")
                continue

            vmin = min(hlist)
            vmax = max(hlist)
            avg = sum(hlist) / len(hlist)
            delta = hlist[-1] - hlist[0]

            if (vmax - vmin) <= self.trend_threshold:
                trend = "stable"
            else:
                if delta > 0:
                    trend = "increasing"
                elif delta < 0:
                    trend = "decreasing"
                else:
                    trend = "stable"

            lines.append(f"  {name}: {trend} (avg {avg:.1f} °C, Δ={delta:+.1f} °C)")

        self.channel_trends_var.set("\n".join(lines))

    def stop_logging(self, error: bool = False):
        if not self.is_logging:
            return

        self.is_logging = False
        self.start_button["state"] = "normal"
        self.stop_button["state"] = "disabled"

        try:
            if self.logger is not None:
                self.logger.close()
        except Exception:
            pass

        self.logger = None

        try:
            if self.csv_file is not None:
                self.csv_file.close()
        except Exception:
            pass

        self.csv_file = None
        self.csv_writer = None

        if not error and self.data_filename and HAVE_OPENPYXL:
            create_colored_excel(self.data_filename)

        if self.arduino is not None:
            try:
                self.arduino.close()
            except Exception:
                pass

            self.arduino = None

        self.set_status("Idle.")

        if not error and self.data_filename:
            messagebox.showinfo("Logging finished", f"Data saved to:\n{self.data_filename}")

    def on_stop(self):
        if self.is_logging:
            self.stop_logging(error=False)

    def on_close(self):
        if self.is_logging:
            if not messagebox.askyesno(
                "Quit",
                "Logging is still running. Stop and exit?"
            ):
                return

            self.stop_logging(error=True)

        if self.graph_window is not None and self.graph_window.winfo_exists():
            try:
                self.graph_window.destroy()
            except Exception:
                pass

            self.graph_window = None

        self.destroy()


if __name__ == "__main__":
    import traceback

    try:
        app = ThermalLoggerApp()
        app.mainloop()
    except Exception:
        traceback.print_exc()
        input("Error occurred, press Enter to exit...")
